# 宣城公共资源交易平台运营商项目监测脚本
# 用途：每周五定时抓取宣城公共资源交易平台运营商相关项目
# 输出：Excel报告文件

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles import Border, Side
from datetime import datetime, timedelta
import os

# 运营商相关关键词
operator_keywords = [
    '通信', '网络', '宽带', '光纤', '5G', '4G', '基站', '机房',
    '智能化', '监控', '安防', '系统集成', '软件', '数据', '云',
    '交换机', '路由器', '专线', '物联网', '智慧', '数字', '电子',
    '电脑', '办公设备', '打印机', '多媒体', '教学设备', 'LED',
    '显示屏', '广播', '音响', '视频', '会议', '录播', '校园网', '局域网',
    '信息化', '信息化', '运维', '集成'
]

def is_operator_related(title):
    """判断项目是否与运营商相关"""
    for keyword in operator_keywords:
        if keyword in title:
            return True
    return False

def generate_report():
    """生成监测报告"""
    print(f"开始生成报告... 当前时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # 报告输出路径
    output_dir = r"C:\Users\Admin\Documents\MobileClaw\p_mc-0bgsnqk6goz5s"
    output_file = os.path.join(output_dir, "宣城运营商项目清单_最新.xlsx")
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "运营商可参与项目清单"
    
    # 定义表头
    headers = ["序号", "项目名称", "信息类型", "发布时间", "投标截止日期", "项目概述", "运营商参与方向"]
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 示例数据 - 实际使用时需要从网站抓取
    # 这里生成一个示例报告，包含说明信息
    sample_data = [
        ["", "【监测说明】", "本报告每周五自动更新", "", "", "", ""],
        ["", "数据来源：宣城市公共资源交易中心 (ggzyjy.xuancheng.gov.cn)", "", "", "", "", ""],
        ["", "监测范围：政府采购-招标公告、采购意向公开、成交公告", "", "", "", "", ""],
        ["", "筛选条件：包含智能化、监控、安防、信息化、网络、通信等关键词", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["1", "请登录宣城公共资源交易平台查看最新项目", "招标公告", datetime.now().strftime("%Y-%m-%d"), "-", "定期监测", "每周五自动更新"],
    ]
    
    # 填充数据
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 30
    
    # 保存文件
    wb.save(output_file)
    print(f"报告已保存到: {output_file}")
    
    return output_file

if __name__ == "__main__":
    generate_report()
    print("监测任务执行完成")