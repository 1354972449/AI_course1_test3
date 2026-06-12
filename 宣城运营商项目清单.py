from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

wb = Workbook()
ws = wb.active
ws.title = "运营商可参与项目清单"

# 定义表头
headers = ["序号", "项目名称", "信息类型", "发布时间", "投标截止日期", "项目概述", "运营商参与方向"]

# 设置表头样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 运营商相关项目数据（2026年5月以来）
projects = [
    # 招标公告 - 运营商可参与
    [1, "旌德县2026年电子监控升级改造设备采购及安装询价公告", "招标公告", "2026-06-10", "2026-06-17", "电子监控设备采购及安装", "安防监控、系统集成"],
    [2, "2026年宣城市独居老人智慧安防关爱项目招标公告", "招标公告", "2026-06-09", "2026-06-30", "智慧安防系统建设", "智慧安防、物联网、监控系统"],
    [3, "郎溪县十字镇初级中学改造提升工程项目智能化工程竞争性磋商公告", "招标公告", "2026-06-09", "2026-06-23", "智能化工程建设", "智能化工程、系统集成、网络布线"],
    [4, "安徽省宁国市安徽材料工程学校（宁国市技工学校）高技能人才培训竞赛基地采购项目招标公告", "招标公告", "2026-06-11", "2026-07-02", "培训基地设备采购", "教学设备、多媒体、网络设备"],
    [5, "2026年应急救援物资采购询价公告", "招标公告", "2026-06-11", "2026-06-18", "应急救援物资采购", "通信设备、应急装备"],
    
    # 采购意向公开 - 潜在项目
    [6, "宣城经济技术开发区管理委员会2026年6月政府采购意向", "采购意向公开", "2026-06-12", "-", "月度采购意向汇总", "关注后续具体采购公告"],
    [7, "宣城市滨湖学校2026年6月至7月政府采购意向", "采购意向公开", "2026-06-12", "-", "学校采购意向", "教育信息化设备"],
    [8, "宁国市公安局2026年6月政府采购意向", "采购意向公开", "2026-06-08", "-", "公安采购意向", "信息化设备、安防监控"],
    [9, "宁国市住房和城乡建设局2026年6月政府采购意向", "采购意向公开", "2026-06-11", "-", "建设局采购意向", "智能化、信息化"],
    
    # 成交公告 - 已完成项目
    [10, "广德经开区管委会办公用房租赁项目成交结果公告", "成交公告", "2026-06-11", "-", "办公用房租赁", "-"],
    [11, "广德市公安局2025年度辅警被装采购项目成交结果公告", "成交公告", "2026-06-10", "-", "警用装备采购", "服装、装备"],
    [12, "玉山取水口、大豪水厂和雁翅站设备更新及运维项目中标结果公告", "成交公告", "2026-06-11", "-", "水厂设备更新", "自动化设备、监控系统"],
    [13, "安徽省宁国市港口中心卫生院救护车及车载急救设备采购项目中标结果公告", "成交公告", "2026-06-11", "-", "医疗设备采购", "车载设备、医疗设备"],
    
    # 更多招标公告
    [14, "2026年郎溪县中等专业学校运动场看台及教学楼维修改造工程项目竞争性磋商公告", "招标公告", "2026-06-11", "2026-06-26", "学校维修改造", "智能化工程、网络布线"],
    [15, "宣州区2026年中小学幼儿园教师和教育管理干部培训项目竞争性磋商公告", "招标公告", "2026-06-11", "2026-06-23", "教师培训服务", "培训平台、在线教育"],
    [16, "2026年泾县政务服务中心综合楼物业服务采购项目竞争性磋商公告", "招标公告", "2026-06-10", "2026-06-22", "物业服务采购", "信息化管理"],
]

# 填充数据
for row_idx, project in enumerate(projects, 2):
    for col_idx, value in enumerate(project, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# 设置列宽
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 25
ws.column_dimensions['G'].width = 30

# 设置边框
from openpyxl.styles import Border, Side
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws.iter_rows(min_row=1, max_row=len(projects)+1, min_col=1, max_col=7):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

# 保存文件
output_path = "C:\\Users\\Admin\\Documents\\MobileClaw\\p_mc-0bgsnqk6goz5s\\宣城运营商项目清单_2026年5月以来.xlsx"
wb.save(output_path)
print(f"Excel文件已保存到: {output_path}")
print(f"共收录 {len(projects)} 个项目")